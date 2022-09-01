import openpyxl
import datetime

from openpyxl import workbook
from openpyxl.utils import formulas
from openpyxl.styles import fonts
from openpyxl.styles import colors
from openpyxl.styles import named_styles

#Import excel
wb = openpyxl.load_workbook('AD.xlsx')

#call active sheet in excel
sheet1 = wb.active
sheet2 = wb['Admins+Webgate']

#sort A to Z
sheet1.auto_filter.ref = "A1:M2000"
sheet2.auto_filter.ref = "A1:M2000"

#freeze panes
#sheet1.freeze_panes = 'A2'
sheet2.freeze_panes = 'A2'

#change sheet title
sheet1.title = "EquitADusers"

#Create new sheet
#wb.create_sheet(index=1, title = 'Generic+Service+Protected')
#wb.create_sheet(index=2, title = 'Admins+Webgate')
#wb.create_sheet(index=3, title = 'VPN')
#wb.create_sheet(index=4, title = 'AIX')
#wb.create_sheet(index=5, title = 'SQL')
#wb.create_sheet(index=6, title = 'UAC')

#Insert Column
sheet1.insert_cols(6)
sheet1['F1'] = 'Account Status'
for x in sheet1['F2:F2000']:
   cell = x[0]
   cell.value = f"=VLOOKUP(G{cell.row},UAC!A$2:B$38,2,FALSE)"


sheet1.insert_cols(8)
sheet1['H1'] = 'Last Logon'
for y in sheet1['H2:H2000']:
   cell = y[0]
   cell.value = f'=IF(I{cell.row}>0,I{cell.row}/(8.64*10^11) - 109205,"")'

for y in sheet1['H2:H2000']:
   cell = y[0]
   cell.number_format = "m/d/yyyy"


sheet1.insert_cols(10)
sheet1['J1'] = 'Pwd Last Changed'
for z in sheet1['J2:J2000']:
   cell = z[0]
   cell.value = f'=IF(K{cell.row}>0,K{cell.row}/(8.64*10^11) - 109205,"")'

for z in sheet1['J2:J2000']:
   cell = z[0]
   cell.number_format = "m/d/yyyy"




for b in sheet2['B3:B2000']:
   cell = b[0]
   cell.value = f"=VLOOKUP(A{cell.row},EquitADusers!B$2:K$5000,2,FALSE)"

for c in sheet2['C3:C2000']:
   cell = c[0]
   cell.value = f"=VLOOKUP(A{cell.row},EquitADusers!B$2:K$5000,3,FALSE)"

for d in sheet2['D3:D2000']:
   cell = d[0]
   cell.value = f"=VLOOKUP(A{cell.row},EquitADusers!B$2:K$5000,4,FALSE)"

for e in sheet2['E3:E2000']:
   cell = e[0]
   cell.value = f"=VLOOKUP(A{cell.row},EquitADusers!B$2:K$5000,5,FALSE)"

for f in sheet2['F3:F2000']:
   cell = f[0]
   cell.value = f"=VLOOKUP(A{cell.row},EquitADusers!B$2:K$5000,6,FALSE)"

for g in sheet2['G3:G2000']:
   cell = g[0]
   cell.value = f"=VLOOKUP(A{cell.row},EquitADusers!B$2:K$5000,7,FALSE)"

for h in sheet2['H3:H2000']:
   cell = h[0]
   cell.value = f"=VLOOKUP(A{cell.row},EquitADusers!B$2:K$5000,8,FALSE)"

for i in sheet2['I3:I2000']:
   cell = i[0]
   cell.value = f"=VLOOKUP(A{cell.row},EquitADusers!B$2:K$5000,9,FALSE)"

for j in sheet2['J3:J2000']:
   cell = j[0]
   cell.value = f"=VLOOKUP(A{cell.row},EquitADusers!B$2:K$5000,10,FALSE)"


for z in sheet2['G2:G2000']:
   cell = z[0]
   cell.number_format = "m/d/yyyy"

for z in sheet2['I2:I2000']:
   cell = z[0]
   cell.number_format = "m/d/yyyy"


#save excel
wb.save('AD_result.xlsx')


print("Complete!")

a = input()